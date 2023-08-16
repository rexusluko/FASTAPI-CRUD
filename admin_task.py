import asyncio
from datetime import timedelta

import httpx
import openpyxl
from celery import Celery

menu_ids: set[str] = set()
submenu_ids: set[str] = set()
submenus: dict[str, str] = {}
dish_ids: set[str] = set()
dishes: dict[str, tuple[str, str]] = {}

base_url = 'http://web:8000/api/v1'

loop = asyncio.get_event_loop()
semaphore = asyncio.Semaphore(1)


async def create_menu(create_data: dict) -> str:
    async with httpx.AsyncClient() as client:
        response = await client.post(f'{base_url}/menus', json=create_data)
        return response.json()['id']


async def update_menu(menu_id: str, update_data: dict) -> None:
    async with httpx.AsyncClient() as client:
        url = f'{base_url}/menus/{menu_id}'
        response = await client.get(url=url)
        current_menu = response.json()
        if current_menu['title'] != update_data['title'] or current_menu['description'] != update_data['description']:
            await client.patch(url=url, json=update_data)


async def delete_menu(menu_id: str) -> None:
    async with httpx.AsyncClient() as client:
        await client.delete(f'{base_url}/menus/{menu_id}')


async def create_submenu(menu_id: str, create_data: dict) -> str:
    async with httpx.AsyncClient() as client:
        response = await client.post(f'{base_url}/menus/{menu_id}/submenus', json=create_data)
        return response.json()['id']


async def update_submenu(menu_id: str, submenu_id: str, update_data: dict) -> None:
    async with httpx.AsyncClient() as client:
        url = f'{base_url}/menus/{menu_id}/submenus/{submenu_id}'
        response = await client.get(url=url)
        current_submenu = response.json()
        if current_submenu['title'] != update_data['title'] or current_submenu['description'] != update_data[
                'description']:
            await client.patch(url=url, json=update_data)


async def delete_submenu(menu_id: str, submenu_id: str) -> None:
    async with httpx.AsyncClient() as client:
        await client.delete(f'{base_url}/menus/{menu_id}/submenus/{submenu_id}')


async def create_dish(menu_id: str, submenu_id: str, create_data: dict) -> str:
    async with httpx.AsyncClient() as client:
        response = await client.post(f'{base_url}/menus/{menu_id}/submenus/{submenu_id}/dishes', json=create_data)
        return response.json()['id']


async def update_dish(menu_id: str, submenu_id: str, dish_id: str, update_data: dict) -> None:
    async with httpx.AsyncClient() as client:
        url = f'{base_url}/menus/{menu_id}/submenus/{submenu_id}/dishes/{dish_id}'
        response = await client.get(url=url)
        current_dish = response.json()
        if current_dish['title'] != update_data['title'] or current_dish['description'] != update_data['description'] or \
                current_dish['price'] != update_data['price']:
            await client.patch(url=url, json=update_data)


async def delete_dish(menu_id: str, submenu_id: str, dish_id: str) -> None:
    async with httpx.AsyncClient() as client:
        await client.delete(f'{base_url}/menus/{menu_id}/submenus/{submenu_id}/dishes/{dish_id}')


async def change_from_excel(filepath: str) -> None:
    global menu_ids, submenu_ids, submenus, dish_ids, dishes
    cur_menu_ids = set()
    cur_submenu_ids = set()
    cur_submenus = {}
    cur_dish_ids = set()
    cur_dishes = {}
    menu_id_column, menu_title_column, menu_description_column = 1, 2, 3
    submenu_id_column, submenu_title_column, submenu_description_column = 2, 3, 4
    dish_id_column, dish_title_column, dish_description_column, dish_price_column = 3, 4, 5, 6
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    row = 1
    while sheet.cell(row=row, column=menu_description_column).value and (
            sheet.cell(row=row, column=submenu_description_column).value is None):  # цикл по меню
        menu_title = sheet.cell(row=row, column=menu_title_column).value
        menu_description = sheet.cell(row=row, column=menu_description_column).value
        menu_data = {'title': menu_title, 'description': menu_description}
        if sheet.cell(row=row, column=menu_id_column).value:
            menu_id = sheet.cell(row=row, column=menu_id_column).value
            async with semaphore:
                await update_menu(menu_id, menu_data)
        else:
            async with semaphore:
                menu_id = await create_menu(menu_data)
            sheet.cell(row=row, column=menu_id_column).value = menu_id
        cur_menu_ids.add(menu_id)
        row += 1
        while sheet.cell(row=row, column=submenu_description_column).value and (
                sheet.cell(row=row, column=dish_description_column).value is None):  # цикл по подменю
            submenu_title = sheet.cell(row=row, column=submenu_title_column).value
            submenu_description = sheet.cell(row=row, column=submenu_description_column).value
            submenu_data = {'title': submenu_title, 'description': submenu_description}
            if sheet.cell(row=row, column=submenu_id_column).value:
                submenu_id = sheet.cell(row=row, column=submenu_id_column).value
                async with semaphore:
                    await update_submenu(menu_id, submenu_id, submenu_data)
            else:
                async with semaphore:
                    submenu_id = await create_submenu(menu_id, submenu_data)
                sheet.cell(row=row, column=submenu_id_column).value = submenu_id
            cur_submenu_ids.add(submenu_id)
            cur_submenus[submenu_id] = menu_id
            row += 1
            while sheet.cell(row=row, column=dish_description_column).value:  # цикл по блюдам
                dish_title = sheet.cell(row=row, column=dish_title_column).value
                dish_description = sheet.cell(row=row, column=dish_description_column).value
                dish_price = str(sheet.cell(row=row, column=dish_price_column).value)
                dish_data = {'title': dish_title, 'description': dish_description, 'price': dish_price}
                if sheet.cell(row=row, column=dish_id_column).value:
                    dish_id = sheet.cell(row=row, column=dish_id_column).value
                    async with semaphore:
                        await update_dish(menu_id, submenu_id, dish_id, dish_data)
                else:
                    async with semaphore:
                        dish_id = await create_dish(menu_id, submenu_id, dish_data)
                    sheet.cell(row=row, column=dish_id_column).value = dish_id
                cur_dish_ids.add(dish_id)
                cur_dishes[dish_id] = (menu_id, submenu_id)
                row += 1

    dish_ids_not_in_excel = dish_ids - cur_dish_ids
    for dish_id in dish_ids_not_in_excel:
        menu_id, submenu_id = dishes[dish_id]
        await delete_dish(menu_id, submenu_id, dish_id)

    submenu_ids_not_in_excel = submenu_ids - cur_submenu_ids
    for submenu_id in submenu_ids_not_in_excel:
        menu_id = submenus[submenu_id]
        await delete_submenu(menu_id, submenu_id)

    menu_ids_not_in_excel = menu_ids - cur_menu_ids
    for menu_id in menu_ids_not_in_excel:
        await delete_menu(menu_id)

    dish_ids, dishes = cur_dish_ids, cur_dishes
    submenu_ids, submenus = cur_submenu_ids, cur_submenus
    menu_ids = cur_menu_ids

    wb.save(filepath)
    wb.close()


celery_app = Celery('admin_task')

celery_app.conf.update(
    broker_url='pyamqp://guest:guest@rabbitmq:5672//',
    result_backend='rpc://',
    task_serializer='json',
    result_serializer='json',
    accept_content=['json'],
    timezone='Europe/Moscow',
    enable_utc=True,
    beat_schedule={
        'main': {
            'task': 'admin_task.main',
            'schedule': timedelta(seconds=15),
        },
    }
)


async def main_async() -> None:
    excel_path = 'admin/Menu.xlsx'
    await change_from_excel(excel_path)


@celery_app.task
def main():
    result = loop.run_until_complete(main_async())
    return result


if __name__ == '__main__':
    asyncio.run(main_async())
